import openpyxl, json

# Leer original de Table 1
wb = openpyxl.load_workbook(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Estado de afiliados al 24.07.2026.xlsx')
ws = wb['Table 1']

afiliados = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    vals = [c.value for c in row]
    if vals[0] is None:
        continue
    nombre = (vals[1] or '').strip()
    uo = (vals[2] or '').strip()
    cargo = (vals[3] or '').strip()
    afiliados.append({'numero': int(vals[0]), 'nombre': nombre, 'uo': uo, 'cargo': cargo})
    # Debug primera fila
    if len(afiliados) == 1:
        print(f"DEBUG fila 1: nombre={nombre}, uo={uo}, cargo={cargo}")

# Cargar codigos
with open(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\sindicato-web\02_utilidades\afiliados_codigos.json', 'r', encoding='utf-8') as f:
    codigos = json.load(f)
codigos_map = {c['numero']: c['codigo'] for c in codigos}

# Crear nuevo Excel
nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = 'Afiliados'

headers = ['N°', 'APELLIDOS Y NOMBRES', 'DNI', 'UO', 'CARGO', 'CODIGO']
header_font = openpyxl.styles.Font(bold=True, size=11)
header_fill = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
thin_border = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
    top=openpyxl.styles.Side(style='thin'),
    bottom=openpyxl.styles.Side(style='thin'),
)

for col, h in enumerate(headers, 1):
    cell = nws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, a in enumerate(afiliados):
    r = i + 2
    nws.cell(row=r, column=1, value=a['numero'])
    nws.cell(row=r, column=2, value=a['nombre'])
    nws.cell(row=r, column=3, value='')  # DNI
    nws.cell(row=r, column=4, value=a['uo'])
    nws.cell(row=r, column=5, value=a['cargo'])  # CARGO REAL
    nws.cell(row=r, column=6, value=codigos_map.get(a['numero'], ''))

nws.column_dimensions['A'].width = 6
nws.column_dimensions['B'].width = 45
nws.column_dimensions['C'].width = 12
nws.column_dimensions['D'].width = 42
nws.column_dimensions['E'].width = 28
nws.column_dimensions['F'].width = 12

nws.auto_filter.ref = f'A1:F{len(afiliados)+1}'
nws.freeze_panes = 'A2'

out = r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Afiliados_SIUTCASJNJ_LIMPIO.xlsx'
nwb.save(out)

# Verificar
wbv = openpyxl.load_workbook(out)
wsv = wbv.active
print(f'\nVerificacion:')
for row in wsv.iter_rows(min_row=2, max_row=5, values_only=True):
    print(f'  {row[0]:2d} | {str(row[1])[:35]:35s} | UO={str(row[3])[:30]:30s} | CARGO={str(row[4])[:25]}')
print(f'\nOK - {out}')
