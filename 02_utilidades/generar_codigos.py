import openpyxl, json, random, string

wb = openpyxl.load_workbook(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Estado de afiliados al 24.07.2026.xlsx')
ws = wb.active

afiliados = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    vals = [c.value for c in row]
    if vals[0] is None:
        continue
    afiliados.append({
        'numero': int(vals[0]),
        'nombre': vals[1].strip() if vals[1] else '',
        'uo': vals[2].strip() if vals[2] else '',
        'cargo': vals[3].strip() if vals[3] else ''
    })

codes = set()
for a in afiliados:
    while True:
        c = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        if c not in codes:
            codes.add(c)
            a['codigo'] = c
            break

print(f'Total afiliados: {len(afiliados)}')
for a in afiliados[:5]:
    print(f"  {a['numero']:2d}. {a['nombre'][:40]:40s} -> {a['codigo']}")

path = r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\sindicato-web\02_utilidades\afiliados_codigos.json'
with open(path, 'w', encoding='utf-8') as f:
    json.dump(afiliados, f, ensure_ascii=False, indent=2)
print(f'\nJSON guardado: {path}')
