import openpyxl, json, re

# 1. Leer hoja CARGOS (data fuente correcta)
wb_orig = openpyxl.load_workbook(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Estado de afiliados al 24.07.2026.xlsx')
ws_orig = wb_orig['CARGOS']

afiliados = []
for row in ws_orig.iter_rows(min_row=3, max_row=ws_orig.max_row):
    vals = [c.value for c in row]
    if vals[0] is None:
        continue
    afiliados.append({
        'numero': int(vals[0]),
        'nombre': (vals[1] or '').strip().upper(),
        'uo': (vals[2] or '').strip(),
        'cargo': (vals[3] or '').strip(),
    })

print(f'{len(afiliados)} afiliados desde hoja CARGOS')

# 2. Leer export con emails y telefonos (si existe)
emails_map = {}
telefonos_map = {}
cumples_map = {}
export_path = r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Afiliados_Exportado_2026-08-06.xlsx'
try:
    wb_exp = openpyxl.load_workbook(export_path)
    ws_exp = wb_exp.active
    for row in ws_exp.iter_rows(min_row=2, max_row=ws_exp.max_row, values_only=True):
        nombre = (row[1] or '').strip().upper()
        if nombre:
            if row[6]: emails_map[nombre] = str(row[6]).strip()
            if row[7]: telefonos_map[nombre] = str(row[7]).strip()
            if row[5]: cumples_map[nombre] = str(row[5]).strip()
    print(f'{len(emails_map)} emails, {len(telefonos_map)} telefonos, {len(cumples_map)} cumpleaños desde export')
except Exception as e:
    print(f'No se pudo leer export: {e}')

# 3. Cargar codigos pre-generados
with open(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\sindicato-web\02_utilidades\afiliados_codigos.json', 'r', encoding='utf-8') as f:
    codigos = json.load(f)
codigos_map = {c['numero']: c['codigo'] for c in codigos}

# 4. Detectar genero por nombre
# Regla: la ultima palabra del nombre (primer nombre) determina el genero
# Termina en A = Mujer, termina en O/E/L/N/R/S/Z = Hombre
# Nombres compuestos con MARIA = Mujer, JOSE = Hombre
def detectar_genero(nombre):
    partes = nombre.split()
    if not partes:
        return 'M'
    nombre_completo = ' '.join(partes)
    # El primer nombre suele ser la ultima palabra (PE: APELLIDO APELLIDO NOMBRE)
    primer_nombre = partes[-1]
    # Si tiene 4+ partes, el nombre esta en las ultimas 2
    if len(partes) >= 4:
        primer_nombre = partes[-1]
    # Si el nombre completo contiene MARIA, es mujer
    if 'MARIA' in nombre_completo:
        return 'F'
    # Nombres masculinos comunes
    if primer_nombre in ('JOSE', 'JUAN', 'CARLOS', 'LUIS', 'PEDRO', 'MARIO',
                         'JORGE', 'MARTIN', 'HENRY', 'RICHARD', 'DIEGO',
                         'WILLIAM', 'SERGIO', 'MANUEL', 'ALY', 'AUGUSTO',
                         'ROGER', 'JORDAN', 'EDWARD', 'SILVERIO', 'JULIO',
                         'OSCAR', 'WELINTONG', 'JESUS', 'ERNESTO', 'PABLO',
                         'GABRIEL', 'FERNANDO', 'ALFREDO', 'ANTONIO',
                         'CHRISTOPER', 'NIELS', 'GERALDO', 'ESTEBAN',
                         'EDSON', 'FELIX', 'CESAR', 'SALVADOR', 'GONZALO',
                         'NEYSSER', 'WILFREDO', 'RAUL', 'VICTOR',
                         'CRISTIAN', 'JIMMY', 'MIGUEL', 'ADRIAN'):
        return 'M'
    # Nombres femeninos comunes
    if primer_nombre in ('KATTY', 'LUCY', 'MARJORIE', 'ANGELA', 'JOSSY',
                         'JACKELINE', 'SILVIA', 'GIANELLA', 'MARIA',
                         'MARSHA', 'ELIZABETH', 'ANDREA', 'KAREN',
                         'PRISCILA', 'GRETA', 'GRECIA', 'MIDORI',
                         'GIOVANNA', 'MIRIAM', 'ANA', 'MARTHA',
                         'JENNIFER', 'LUISA', 'GUILIANA', 'KATHERINE',
                         'GISSELA', 'KAROL', 'MARGARETH', 'NATHALI',
                         'ROSSY', 'DELIA', 'ANSELMA', 'NOEMI', 'JANE',
                         'MERCEDES', 'JASQUELINE', 'CANDY', 'TANIA',
                         'DOLORES', 'MARGARITA', 'BERTHA', 'ELENA',
                         'PAOLA', 'GRACIELA', 'LIZBETH'):
        return 'F'
    # Heuristica por terminacion
    if primer_nombre.endswith('A'):
        return 'F'
    return 'M'

# 5. Crear Excel final
nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = 'Afiliados'

headers = ['N°', 'APELLIDOS Y NOMBRES', 'DNI', 'UO', 'CARGO', 'CODIGO',
           'CUMPLEAÑOS', 'EMAIL', 'TELEFONO', 'GENERO']
hfont = openpyxl.styles.Font(bold=True, size=11)
hfill = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
thin = openpyxl.styles.Side(style='thin')
border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = nws.cell(row=1, column=col, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.border = border

for i, a in enumerate(afiliados):
    r = i + 2
    nws.cell(row=r, column=1, value=a['numero'])
    nws.cell(row=r, column=2, value=a['nombre'])
    nws.cell(row=r, column=3, value='')  # DNI
    nws.cell(row=r, column=4, value=a['uo'])
    nws.cell(row=r, column=5, value=a['cargo'])
    nws.cell(row=r, column=6, value=codigos_map.get(a['numero'], ''))
    nws.cell(row=r, column=7, value=cumples_map.get(a['nombre'], ''))
    nws.cell(row=r, column=8, value=emails_map.get(a['nombre'], ''))
    nws.cell(row=r, column=9, value=telefonos_map.get(a['nombre'], ''))
    genero = detectar_genero(a['nombre'])
    nws.cell(row=r, column=10, value=genero)

# Anchura columnas
widths = [6, 45, 12, 42, 28, 12, 14, 35, 15, 8]
for i, w in enumerate(widths, 1):
    nws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

nws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(afiliados)+1}'
nws.freeze_panes = 'A2'

out = r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Afiliados_SIUTCASJNJ_LIMPIO.xlsx'
nwb.save(out)

# Resumen
hombres = sum(1 for a in afiliados if detectar_genero(a['nombre']) == 'M')
mujeres = sum(1 for a in afiliados if detectar_genero(a['nombre']) == 'F')
print(f'\nGenero: {hombres} Hombres, {mujeres} Mujeres')
print(f'Con cumpleaños: {len(cumples_map)}')
print(f'Con email: {len(emails_map)}')
print(f'Con telefono: {len(telefonos_map)}')
print(f'\nArchivo: {out}')

# Mostrar muestra
print('\nMuestra de generos detectados:')
muestras = [a for a in afiliados if detectar_genero(a['nombre']) == 'F'][:5]
muestras += [a for a in afiliados if detectar_genero(a['nombre']) == 'M'][:5]
for a in muestras:
    g = detectar_genero(a['nombre'])
    print(f'  {g} | {a["nombre"][:40]}')
