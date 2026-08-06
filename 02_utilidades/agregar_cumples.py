import openpyxl, json

# Regenerar desde CARGOS
wb = openpyxl.load_workbook(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Estado de afiliados al 24.07.2026.xlsx')
ws = wb['CARGOS']

afiliados = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    vals = [c.value for c in row]
    if vals[0] is None:
        continue
    afiliados.append({
        'numero': int(vals[0]),
        'nombre': (vals[1] or '').strip().upper(),
        'uo': (vals[2] or '').strip(),
        'cargo': (vals[3] or '').strip()
    })

# Cargar codigos
with open(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\sindicato-web\02_utilidades\afiliados_codigos.json', 'r', encoding='utf-8') as f:
    codigos = json.load(f)
codigos_map = {c['numero']: c['codigo'] for c in codigos}

# Cumpleaños
cumples = {
    'VALLEJOS HERENCIA MARIO ALBERTO': '07/08/1981',
    'CRUZ MANCHEGO MARSHA JIMENA': '12/08/1989',
    'FALCONI PEÑA ANDREA PATRICIA': '13/08/1985',
    'SILVA MACETAS GISSELA MARGARITA': '19/08/1981',
    'PALMADERA ARROSTICO ERMELINDO JULIO': '24/08/1987',
    'CHOQUE AVALOS ROSSY MARGARITA': '26/08/1996',
    'ORDOÑEZ ESCARZA PAVEL': '28/08/1978',
    'HINOSTROZA MARINGO CARLOS GABRIEL': '30/08/1975',
    'FLORES MARQUEZ PRISCILA RAYZA': '31/08/1992',
}

def normalizar(s):
    import re
    s = s.upper().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').replace('Ñ','N')
    return re.sub(r'\s+', ' ', s).strip()

cumples_norm = {normalizar(k): v for k, v in cumples.items()}

# Crear Excel
nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = 'Afiliados'

headers = ['N°', 'APELLIDOS Y NOMBRES', 'DNI', 'UO', 'CARGO', 'CODIGO', 'CUMPLEAÑOS']
hfont = openpyxl.styles.Font(bold=True, size=11)
hfill = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
thin = openpyxl.styles.Side(style='thin')
border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = nws.cell(row=1, column=col, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.border = border

encontrados = 0
for i, a in enumerate(afiliados):
    r = i + 2
    nws.cell(row=r, column=1, value=a['numero'])
    nws.cell(row=r, column=2, value=a['nombre'])
    nws.cell(row=r, column=3, value='')
    nws.cell(row=r, column=4, value=a['uo'])
    nws.cell(row=r, column=5, value=a['cargo'])
    nws.cell(row=r, column=6, value=codigos_map.get(a['numero'], ''))

    nom_norm = normalizar(a['nombre'])
    for cn, cv in cumples_norm.items():
        if cn in nom_norm:
            nws.cell(row=r, column=7, value=cv)
            encontrados += 1
            print(f"  #{a['numero']:2d} {cv} -> {a['nombre'][:45]}")
            break

nws.column_dimensions['A'].width = 6
nws.column_dimensions['B'].width = 45
nws.column_dimensions['C'].width = 12
nws.column_dimensions['D'].width = 42
nws.column_dimensions['E'].width = 28
nws.column_dimensions['F'].width = 12
nws.column_dimensions['G'].width = 14
nws.auto_filter.ref = f'A1:G{len(afiliados)+1}'
nws.freeze_panes = 'A2'

out = r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Afiliados_SIUTCASJNJ_LIMPIO.xlsx'
nwb.save(out)
print(f'\n{encontrados}/9 cumpleanos agregados.')
print(f'Archivo: {out}')
