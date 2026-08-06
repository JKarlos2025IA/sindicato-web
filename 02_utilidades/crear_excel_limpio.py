import openpyxl, json

wb = openpyxl.load_workbook(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Estado de afiliados al 24.07.2026.xlsx')
ws = wb['Table 1']

afiliados = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    vals = [c.value for c in row]
    if vals[0] is None:
        continue
    afiliados.append({
        'numero': int(vals[0]),
        'nombre': vals[1].strip() if vals[1] else '',
        'uo': vals[2].strip() if vals[2] else '',
        'cargo': vals[3].strip() if vals[3] else '',
    })

with open(r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\sindicato-web\02_utilidades\afiliados_codigos.json', 'r', encoding='utf-8') as f:
    codigos = json.load(f)

codigos_map = {c['numero']: c['codigo'] for c in codigos}

nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = 'Afiliados'

headers = ['N°', 'APELLIDOS Y NOMBRES', 'DNI', 'UO', 'CARGO', 'CODIGO']
for col, h in enumerate(headers, 1):
    cell = nws.cell(row=1, column=col, value=h)
    cell.font = openpyxl.styles.Font(bold=True, size=11)
    cell.fill = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    cell.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

for i, a in enumerate(afiliados):
    row = i + 2
    nws.cell(row=row, column=1, value=a['numero'])
    nws.cell(row=row, column=2, value=a['nombre'])
    nws.cell(row=row, column=3, value='')  # DNI vacio
    nws.cell(row=row, column=4, value=a['uo'])
    nws.cell(row=row, column=5, value=a['cargo'])
    nws.cell(row=row, column=6, value=codigos_map.get(a['numero'], ''))

# Ajustar anchos
nws.column_dimensions['A'].width = 6
nws.column_dimensions['B'].width = 45
nws.column_dimensions['C'].width = 12
nws.column_dimensions['D'].width = 42
nws.column_dimensions['E'].width = 25
nws.column_dimensions['F'].width = 12

# Filtro y congelar
nws.auto_filter.ref = f'A1:F{len(afiliados)+1}'
nws.freeze_panes = 'A2'

out = r'C:\Users\juan.montenegro\Desktop\PAGINA WEB-SINDICATO\Afiliados_SIUTCASJNJ_LIMPIO.xlsx'
nwb.save(out)
print(f'Excel limpio creado: {out}')
print(f'{len(afiliados)} afiliados con cabeceras: N° | APELLIDOS Y NOMBRES | DNI | UO | CARGO | CODIGO')
